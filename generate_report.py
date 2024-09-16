from docx import Document
from docx.shared import Inches
import pandas as pd
import subprocess
from docx.enum.dml import MSO_THEME_COLOR_INDEX
import docx
import os
from google.cloud import bigquery
import json
from google.cloud import storage
from datetime import datetime
from time import time, sleep
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def get_filename_without_extension(file_name):
    """
    Obtiene el nombre del archivo sin su extensión.
    
    :param file_name: Nombre del archivo con extensión.
    :return: Nombre del archivo sin extensión.
    """
    extension = file_name.split(".")[-1]
    filename_without_extension = file_name[:-len(extension) - 1]
    return filename_without_extension


# Instanciar un cliente de Google Cloud Storage
storage_client = storage.Client()

# Instanciar un cliente de BigQuery
bqclient = bigquery.Client()

# Consulta SQL para BigQuery que selecciona registros sin datetime_output
query = """
    SELECT *
    FROM `lib-ia.d_client_info.f_request_stage1_view`
    WHERE datetime_output IS NULL
"""

# Ejecuta la consulta y convierte el resultado en un DataFrame
query_job = bqclient.query(query).to_dataframe()

# Iterar sobre cada registro del DataFrame
for index, registro in query_job.iterrows():
    print(registro.text_name)
    print(registro.uuid)

    # Procesar el resumen: capitalizar las frases
    input_resumen = registro.resume
    sentences_resumen = input_resumen.split(". ")
    capitalized_sentences = [sentence.capitalize() for sentence in sentences_resumen]
    input_resumen = ". ".join(capitalized_sentences)
    input_libro = registro.text_name

    # Procesar los tópicos (categoría): normalizar el JSON a DataFrame
    input_categoria = registro.topics
    json_obj = json.loads(input_categoria)
    input_categoria = pd.json_normalize(json_obj['topics'], sep='_')

    # Procesar los personajes: convertir el JSON a DataFrame
    input_personajes = registro.characters
    json_obj_personajes = json.loads(input_personajes)
    df_personajes = pd.DataFrame(json_obj_personajes)

    # Crear un DataFrame para los campos generales del registro
    data_generales = pd.DataFrame({
        'nombre_archivo': [registro.text_name],
        'cantidad_palabras': [registro.word_count],
        'caracteres_sin_espacios': [registro.char_without_spaces],
        'caracteres_con_espacios': [registro.char_with_spaces],
        'cantidad_imagenes': [registro.img_amount],
        'publico_potencial': [registro.potential_audience],
        'resumen': [input_resumen],
        'fecha_hora_pedido': [registro.datetime_request],
        'usuario': [registro.desc_user],
        'uuid': [registro.uuid]
    })

    # Crear una nueva instancia de Workbook para el archivo Excel
    workbook = Workbook()

    # Agregar el DataFrame de campos generales a una hoja llamada 'general'
    worksheet_general = workbook.create_sheet(title='general')
    for row in dataframe_to_rows(data_generales, index=False, header=True):
        worksheet_general.append(row)

    # Agregar los tópicos a una hoja llamada 'topicos'
    worksheet_topicos = workbook.create_sheet(title='topicos')
    for row in dataframe_to_rows(input_categoria, index=False, header=True):
        worksheet_topicos.append(row)

    # Agregar los personajes a una hoja llamada 'personajes'
    worksheet_personajes = workbook.create_sheet(title='personajes')
    for row in dataframe_to_rows(df_personajes, index=False, header=True):
        worksheet_personajes.append(row)

    # Eliminar la hoja predeterminada "Sheet"
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)

    # Ajustar los anchos de columna para todas las hojas del Excel (opcional)
    for sheetname in workbook.sheetnames:
        worksheet = workbook[sheetname]
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Guardar el archivo Excel en el directorio de salida
    excel_file = 'lib-ia/output/' + get_filename_without_extension(input_libro) + '.xlsx'
    workbook.save(excel_file)

    # Guardar el archivo en Cloud Storage
    client = storage.Client()
    bucket = client.get_bucket('lib-ia')
    blob = bucket.blob('reportes/created_stage1/' + get_filename_without_extension(input_libro) + '.xlsx')
    
    # Subir el archivo al bucket de Google Cloud Storage
    blob.upload_from_filename(excel_file)

    # Crear el URI del archivo subido
    uri = 'reportes/created_stage1/' + get_filename_without_extension(input_libro) + '.xlsx'
    print(uri)

    # Insertar los detalles del archivo generado en BigQuery
    table_id = "lib-ia.d_client_info.f_request_output"
    rows_to_insert = [
        {
            'uuid': registro.uuid,
            'datetime_output': str(datetime.now()),
            'output_type': 'XLSX',
            'uri_output': uri,
            'stage': '1'
        }
    ]

    # Insertar las filas en la tabla de BigQuery
    errors = bqclient.insert_rows_json(table_id, rows_to_insert)
    
    if errors == []:
        print("New rows have been added.")
    else:
        print(f"Encountered errors while inserting rows: {errors}")
